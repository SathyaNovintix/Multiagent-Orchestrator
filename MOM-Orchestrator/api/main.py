"""
FastAPI App — AgentMesh AI
"""
from __future__ import annotations
from contextlib import asynccontextmanager
from dotenv import load_dotenv

load_dotenv()

import io
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from orchestrator.registry import build_registry
from orchestrator.core import init_orchestrator, new_session, run_pipeline
from storage.mongo_client import (
    ensure_indexes, ping as mongo_ping,
    get_mom, get_messages, save_message, list_sessions, get_session,
)
from pdf.generator import generate_pdf_bytes
from formats.registry import get_all_formats, get_all_formats_async, get_format, register_custom_format, get_default_format
from formats.template_parser import parse_template_file


# ---------------------------------------------------------------------------
# Lifespan
# ---------------------------------------------------------------------------

@asynccontextmanager
async def lifespan(app: FastAPI):
    await ensure_indexes()
    build_registry()
    init_orchestrator()
    yield


app = FastAPI(title="AgentMesh AI — MOM Orchestrator", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Agent Testing Routers
# ---------------------------------------------------------------------------

from agents.conversational.router import router as conversational_router
from agents.intent_refiner.router import router as intent_refiner_router
from agents.language_detector.router import router as language_detector_router
from agents.translator.router import router as translator_router
from agents.topic_extractor.router import router as topic_extractor_router
from agents.decision_extractor.router import router as decision_extractor_router
from agents.action_extractor.router import router as action_extractor_router
from agents.formatter.router import router as formatter_router
from agents.response_generator.router import router as response_generator_router
from agents.speech_to_text.router import router as speech_to_text_router
from tools.microsoft_teams.router import router as teams_router
from api.task_assignment import router as task_assignment_router

app.include_router(conversational_router)
app.include_router(intent_refiner_router)
app.include_router(language_detector_router)
app.include_router(translator_router)
app.include_router(topic_extractor_router)
app.include_router(decision_extractor_router)
app.include_router(action_extractor_router)
app.include_router(formatter_router)
app.include_router(response_generator_router)
app.include_router(speech_to_text_router)
app.include_router(teams_router)
app.include_router(task_assignment_router)


# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------

class NewSessionRequest(BaseModel):
    label: Optional[str] = None


class NewSessionResponse(BaseModel):
    session_id: str
    label: str
    created_at: str


class RunRequest(BaseModel):
    session_id: str
    input_type: str
    content: str
    language_hint: Optional[str] = None
    intent: str = "auto_detect"
    format_id: str = "standard"


class RunResponse(BaseModel):
    type: str
    user_message: Optional[str] = None
    file_url: Optional[str] = None
    structured_mom: Optional[dict] = None
    message: Optional[str] = None
    prompt: Optional[str] = None
    trace: Optional[list] = None


class MessageIn(BaseModel):
    id: str
    role: str
    content: str
    type: Optional[str] = "text"
    mom: Optional[dict] = None
    file_url: Optional[str] = None
    timestamp: str


# ---------------------------------------------------------------------------
# Sessions
# ---------------------------------------------------------------------------

@app.post("/session", response_model=NewSessionResponse)
async def create_session_endpoint(body: NewSessionRequest = NewSessionRequest()):
    label = body.label or f"Session {datetime.utcnow().strftime('%d %b %H:%M')}"
    session = await new_session(label=label)
    return NewSessionResponse(
        session_id=session.session_id,
        label=session.label,
        created_at=session.created_at,
    )


@app.get("/sessions")
async def get_all_sessions():
    """Returns all sessions (newest first) for sidebar restore."""
    sessions = await list_sessions(limit=100)
    return {"sessions": sessions}


@app.get("/sessions/{session_id}")
async def get_one_session(session_id: str):
    session = await get_session(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found.")
    return session.model_dump()


# ---------------------------------------------------------------------------
# Messages
# ---------------------------------------------------------------------------

@app.get("/sessions/{session_id}/messages")
async def get_session_messages(session_id: str):
    msgs = await get_messages(session_id)
    return {"messages": msgs}


@app.post("/sessions/{session_id}/messages")
async def post_message(session_id: str, msg: MessageIn):
    await save_message(session_id, msg.model_dump())
    return {"ok": True}


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

@app.post("/run", response_model=RunResponse)
async def run(request: RunRequest):
    result = await run_pipeline(
        session_id=request.session_id,
        input_type=request.input_type,
        content=request.content,
        language_hint=request.language_hint,
        intent=request.intent,
        format_id=request.format_id,
    )
    return RunResponse(**result)


# ---------------------------------------------------------------------------
# MOM download — PDF
# ---------------------------------------------------------------------------

@app.get("/api/mom/{mom_id}/download")
async def download_mom_pdf(mom_id: str, format_id: str = "standard"):
    mom = await get_mom(mom_id)
    if mom is None:
        raise HTTPException(status_code=404, detail="MOM not found.")
    fmt = get_format(format_id) or get_format("standard")
    pdf_bytes = await generate_pdf_bytes(mom, title=fmt.get("name", "Minutes of Meeting"), fmt=fmt)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=MOM_{mom_id[:8]}_{format_id}.pdf"},
    )


# ---------------------------------------------------------------------------
# MOM download — Excel
# ---------------------------------------------------------------------------

@app.get("/api/mom/{mom_id}/download/excel")
async def download_mom_excel(mom_id: str, format_id: str = "standard"):
    """
    Generate and download an Excel (.xlsx) file of the MOM.
    For custom templates: renders each section's fields as a separate sheet.
    For standard formats: creates Topics / Decisions / Actions sheets.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    mom = await get_mom(mom_id)
    if mom is None:
        raise HTTPException(status_code=404, detail="MOM not found.")

    fmt = get_format(format_id) or get_format("standard")
    
    # Prefer template_structure from the MOM document itself (what the AI actually used)
    # Fallback to the current format definition if not in DB
    template_structure = getattr(mom, "template_structure", None) or (fmt or {}).get("template_structure")
    sections_data = getattr(mom, "sections", None)
    
    accent_hex = (fmt or {}).get("accent_color", "#1a1a2e").lstrip("#")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    def _safe_str(val) -> str:
        if val in (None, "", "None", "N/A", [], {}):
            return "N/A"
        if isinstance(val, (str, int, float, bool)):
            return str(val).strip() if str(val).strip() else "N/A"
        if isinstance(val, list):
            parts = []
            for item in val:
                if isinstance(item, dict):
                    parts.append(", ".join(f"{k}: {v}" for k, v in item.items() if not isinstance(v, (dict, list))))
                else:
                    parts.append(str(item))
            return " | ".join(parts) if parts else "N/A"
        if isinstance(val, dict):
            for key in ("value", "text", "content", "data"):
                if key in val and not isinstance(val[key], (dict, list)):
                    return str(val[key]) if val[key] is not None else "N/A"
            parts = [f"{k}: {v}" for k, v in val.items() if not isinstance(v, (dict, list)) and v]
            return ", ".join(parts) if parts else "N/A"
        return str(val)

    def _header_style(cell, hex_color: str = "1A1A2E"):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 60)

    # ── Single sheet configuration ──
    ws = wb.create_sheet(title="MOM Report")
    curr_row = 1

    def _append_section_header(title: str, hex_color: str = "1A1A2E"):
        nonlocal curr_row
        if curr_row > 1:
            curr_row += 1
        cell = ws.cell(curr_row, 1, title)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1

    def _append_table_header(headers: list[str], hex_color: str = "4A5568"):
        nonlocal curr_row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(curr_row, col_idx, h)
            _header_style(cell, hex_color)
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

    # ── Custom sections ──
    if template_structure and template_structure.get("sections"):
        mom_sections: dict = getattr(mom, "sections", None) or {}
        fields_meta: dict = template_structure.get("fields", {})

        for section in template_structure["sections"]:
            sid = section.get("id", "section")
            slabel = section.get("label", sid)
            section_data = mom_sections.get(sid)
            fields = fields_meta.get(sid, [])

            _append_section_header(slabel, accent_hex)

            if not section_data or section_data == "N/A":
                ws.cell(curr_row, 1, "N/A").font = Font(italic=True)
                curr_row += 1
                continue

            if isinstance(section_data, list):
                # Table-like section
                if section_data and isinstance(section_data[0], dict):
                    headers = [f.get("label") or f.get("name") for f in fields] if fields else list(section_data[0].keys())
                    _append_table_header(headers)
                    for item in section_data:
                        if isinstance(item, dict):
                            row_vals = []
                            if fields:
                                for f in fields:
                                    fid = f.get("id") or f.get("name","").lower().replace(" ","_")
                                    row_vals.append(_safe_str(item.get(fid)))
                            else:
                                row_vals = [_safe_str(v) for v in item.values()]
                            ws.append(row_vals)
                            curr_row += 1
                else:
                    for item in section_data:
                        ws.cell(curr_row, 1, _safe_str(item))
                        curr_row += 1
            elif isinstance(section_data, dict):
                # Key-Value section
                for k, v in section_data.items():
                    label = k.replace("_", " ").title()
                    # Try to find label in fields
                    if fields:
                        for f in fields:
                            fid = f.get("id") or f.get("name","").lower().replace(" ","_")
                            if fid == k:
                                label = f.get("label") or f.get("name") or label
                                break
                    ws.cell(curr_row, 1, label).font = Font(bold=True)
                    ws.cell(curr_row, 2, _safe_str(v))
                    curr_row += 1
            else:
                ws.cell(curr_row, 1, _safe_str(section_data))
                curr_row += 1

    # ── Standard sections ──
    # Append topics
    _append_section_header("Discussion Topics", "2B6CB0")
    if mom.topics:
        _append_table_header(["#", "Title", "Summary", "Timestamp"])
        for i, t in enumerate(mom.topics, 1):
            ws.append([i, _safe_str(t.title), _safe_str(t.summary), _safe_str(getattr(t, "timestamp", ""))])
            curr_row += 1
    else:
        ws.cell(curr_row, 1, "N/A").font = Font(italic=True)
        curr_row += 1

    # Append decisions
    _append_section_header("Decisions Made", "2F855A")
    if mom.decisions:
        _append_table_header(["#", "Decision", "Owner", "Condition"])
        for i, d in enumerate(mom.decisions, 1):
            ws.append([i, _safe_str(d.decision), _safe_str(d.owner), _safe_str(getattr(d, "condition", ""))])
            curr_row += 1
    else:
        ws.cell(curr_row, 1, "N/A").font = Font(italic=True)
        curr_row += 1

    # Append actions
    _append_section_header("Action Items", "C05621")
    if mom.actions:
        _append_table_header(["#", "Task", "Owner", "Deadline", "Priority"])
        for i, a in enumerate(mom.actions, 1):
            ws.append([i, _safe_str(a.task), _safe_str(a.owner), _safe_str(getattr(a, "deadline", "")), _safe_str(a.priority)])
            curr_row += 1
    else:
        ws.cell(curr_row, 1, "N/A").font = Font(italic=True)
        curr_row += 1

    _auto_width(ws)

    # Fallback: at least one sheet
    if not wb.sheetnames or (len(wb.sheetnames) == 1 and ws.max_row == 0):
        if "MOM Report" not in wb.sheetnames:
            ws = wb.create_sheet("MOM Report")
        ws.append(["No data available"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=MOM_{mom_id[:8]}_{format_id}.xlsx"},
    )






# ---------------------------------------------------------------------------
# Formats
# ---------------------------------------------------------------------------

@app.get("/formats")
async def list_formats():
    formats = await get_all_formats_async()
    return {"formats": formats}


@app.post("/formats/preview")
async def preview_template(
    template_file: UploadFile = File(...),
):
    """
    Preview the structure of an uploaded template file without saving it.
    Returns the parsed sections and fields.
    """
    print(f"[preview] Received file: {template_file.filename}")
    try:
        file_bytes = await template_file.read()
        print(f"[preview] File size: {len(file_bytes)} bytes")
        template_structure = parse_template_file(file_bytes, template_file.filename)
        print(f"[preview] Parsed structure: {template_structure}")
        return template_structure
    except Exception as e:
        print(f"[preview] Error parsing template: {e}")
        import traceback
        traceback.print_exc()
        return {
            "sections": [],
            "fields": {},
            "source": "error",
            "filename": template_file.filename,
            "parse_error": str(e)
        }


@app.post("/formats/custom")
async def upload_custom_format(
    name: str = Form(...),
    description: str = Form(""),
    accent_color: str = Form("#e6a817"),
    header_color: str = Form("#1a1a2e"),
    sections: str = Form("topics,decisions,actions"),
    template_file: Optional[UploadFile] = File(None),
):
    """
    Create a custom MOM format with optional template file upload.
    If a template file is provided, its structure will be parsed and used.
    """
    template_structure = None
    custom_labels = {}
    
    # Parse uploaded template file if provided
    if template_file:
        try:
            file_bytes = await template_file.read()
            template_structure = parse_template_file(file_bytes, template_file.filename)
            
            # Extract sections and labels from parsed template
            if template_structure.get('sections'):
                sections = ",".join([s['id'] for s in template_structure['sections']])
                # Build custom labels from template
                for section in template_structure['sections']:
                    if section['label'] != section['id'].capitalize():
                        custom_labels[section['id']] = section['label']
        except Exception as e:
            # If parsing fails, continue with default sections
            print(f"[formats] Template parsing failed: {e}")
    
    fmt = register_custom_format({
        "name": name,
        "description": description,
        "accent_color": accent_color,
        "header_color": header_color,
        "sections": [s.strip() for s in sections.split(",")],
        "custom_labels": custom_labels if custom_labels else None,
        "icon": "📄",
        "template_structure": template_structure,
    })
    return {"format": fmt}


@app.post("/upload-audio")
async def upload_audio_file(audio_file: UploadFile = File(...)):
    """
    Upload audio file and save to MongoDB GridFS.
    Returns file_id for later retrieval and transcription.
    """
    from storage.mongo_client import save_audio_file

    try:
        # Read file content
        file_bytes = await audio_file.read()

        # Detect content type
        filename = audio_file.filename or "audio.mp3"
        content_type = audio_file.content_type or "audio/mpeg"

        # Save to GridFS
        file_id = await save_audio_file(file_bytes, filename, content_type)

        return {
            "file_path": file_id,  # Return GridFS file_id as "file_path"
            "filename": filename,
            "size": len(file_bytes)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

@app.get("/health")
async def health():
    mongo_ok = await mongo_ping()
    return {
        "status": "ok" if mongo_ok else "degraded",
        "mongodb": "connected" if mongo_ok else "disconnected",
    }


# ---------------------------------------------------------------------------
# MOM retrieval and update endpoints
# ---------------------------------------------------------------------------

@app.get("/api/mom/{mom_id}")
async def get_mom_endpoint(mom_id: str):
    """
    Get a MOM document by ID.
    """
    mom = await get_mom(mom_id)
    if mom is None:
        raise HTTPException(status_code=404, detail="MOM not found")
    
    return mom.model_dump()


class UpdateMOMRequest(BaseModel):
    topics: Optional[list] = None
    decisions: Optional[list] = None
    actions: Optional[list] = None
    sections: Optional[dict] = None
    participants: Optional[list] = None


@app.put("/api/mom/{mom_id}")
async def update_mom(mom_id: str, request: UpdateMOMRequest):
    """
    Update an existing MOM document.
    Only updates the fields that are provided in the request.
    """
    from storage.mongo_client import save_mom
    from schemas.contracts import Topic, Decision, Action
    
    print(f"[UPDATE MOM] Received update request for MOM ID: {mom_id}")
    print(f"[UPDATE MOM] Request data: topics={request.topics is not None}, decisions={request.decisions is not None}, actions={request.actions is not None}, sections={request.sections is not None}")
    
    # Fetch existing MOM
    mom = await get_mom(mom_id)
    if mom is None:
        print(f"[UPDATE MOM] ERROR: MOM not found: {mom_id}")
        raise HTTPException(status_code=404, detail="MOM not found")
    
    print(f"[UPDATE MOM] Found existing MOM: {mom_id}")
    
    # Update fields if provided
    if request.topics is not None:
        print(f"[UPDATE MOM] Updating {len(request.topics)} topics")
        mom.topics = [Topic(**t) if isinstance(t, dict) else t for t in request.topics]
    
    if request.decisions is not None:
        print(f"[UPDATE MOM] Updating {len(request.decisions)} decisions")
        mom.decisions = [Decision(**d) if isinstance(d, dict) else d for d in request.decisions]
    
    if request.actions is not None:
        print(f"[UPDATE MOM] Updating {len(request.actions)} actions")
        mom.actions = [Action(**a) if isinstance(a, dict) else a for a in request.actions]
    
    if request.sections is not None:
        print(f"[UPDATE MOM] Updating sections: {list(request.sections.keys())}")
        mom.sections = request.sections
    
    if request.participants is not None:
        print(f"[UPDATE MOM] Updating {len(request.participants)} participants")
        mom.participants = request.participants
    
    # Save updated MOM
    print(f"[UPDATE MOM] Saving to database...")
    await save_mom(mom)
    print(f"[UPDATE MOM] Successfully saved MOM: {mom_id}")
    
    return {
        "status": "success",
        "message": "MOM updated successfully",
        "mom_id": mom_id
    }
